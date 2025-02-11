AIzaSyC25g1ANIYDSbbjtXfEDpGAt7unZoZsPC0
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def analyze_questions(input_csv, output_xlsx):
    """
    Reads a CSV file, analyzes question-answer pairs based on predefined rules, applies color formatting, and saves to an Excel file.
    """
    # Load CSV into DataFrame
    df = pd.read_csv(input_csv, encoding="utf-8-sig")
    
    # Standardize column names
    df.columns = df.columns.str.strip().str.lower()
    
    # Ensure 'question' and 'answer' columns exist
    if "question" not in df.columns or "answer" not in df.columns:
        raise KeyError("CSV file must contain 'question' and 'answer' columns")
    
    # Create 'status' column for Pass/Fail
    df["status"] = ""

    # Define ranges and conditions
    for index, row in df.iterrows():
        question_number = index + 1  # Adjusting for 1-based index
        
        if 2 <= question_number <= 16:
            df.at[index, "status"] = "Fail - Non-English Question"
        elif 17 <= question_number <= 49:
            if row["answer"].strip():  # If an answer is present, it's an error
                df.at[index, "status"] = "Fail - Answer should not be generated"
            else:
                df.at[index, "status"] = "Pass"
        elif 50 <= question_number <= 64:
            if row["answer"].strip():
                df.at[index, "status"] = "Pass - Complex Question"
            else:
                df.at[index, "status"] = "Fail - No answer for complex question"
        elif 65 <= question_number <= 205:
            if row["answer"].strip():
                df.at[index, "status"] = "Pass"
            else:
                df.at[index, "status"] = "Fail - No answer provided"

    # Save to Excel with colors
    df.to_excel(output_xlsx, index=False, engine="openpyxl")

    # Load workbook for formatting
    wb = load_workbook(output_xlsx)
    ws = wb.active

    # Define color styles
    section_colors = {
        "Non-English": "FFC7CE",  # Light Red
        "No Answer Allowed": "FFEB9C",  # Yellow
        "Complex Question": "C6E0B4",  # Light Green
        "Valid Answers": "D9E1F2"  # Light Blue
    }
    
    pass_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    fail_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

    # Identify column indexes
    status_col_idx = df.columns.get_loc("status") + 1  # Convert to Excel column index

    # Apply row-wise section coloring
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 to skip headers
        question_number = row_idx - 1  # Adjusting for 1-based indexing
        
        # Determine section color
        if 2 <= question_number <= 16:
            section_fill = PatternFill(start_color=section_colors["Non-English"], end_color=section_colors["Non-English"], fill_type="solid")
        elif 17 <= question_number <= 49:
            section_fill = PatternFill(start_color=section_colors["No Answer Allowed"], end_color=section_colors["No Answer Allowed"], fill_type="solid")
        elif 50 <= question_number <= 64:
            section_fill = PatternFill(start_color=section_colors["Complex Question"], end_color=section_colors["Complex Question"], fill_type="solid")
        elif 65 <= question_number <= 205:
            section_fill = PatternFill(start_color=section_colors["Valid Answers"], end_color=section_colors["Valid Answers"], fill_type="solid")
        else:
            section_fill = None  # No color for rows outside defined range

        # Apply section color
        if section_fill:
            for col_idx in range(1, ws.max_column + 1):  # Apply to all columns
                ws.cell(row=row_idx, column=col_idx).fill = section_fill
        
        # Apply Pass/Fail color
        status_cell = ws.cell(row=row_idx, column=status_col_idx)
        if "Fail" in status_cell.value:
            status_cell.fill = fail_color
        elif "Pass" in status_cell.value:
            status_cell.fill = pass_color

    # Save the formatted Excel file
    wb.save(output_xlsx)
    print(f"✅ Analysis complete! Colored report saved to '{output_xlsx}'")

# Example Usage:
# analyze_questions("Updated_Test_Data.csv", "Analysis_Report.xlsx")
